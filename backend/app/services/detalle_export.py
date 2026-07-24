"""Genera el Excel de descarga rellenando la plantilla del usuario
(app/resources/plantilla.xlsx), conservando exactamente sus colores, layout,
secciones fijas (Agentes de Aduanas, Pagos al Personal, Seguros) y fórmulas.

Solo se inyectan los datos en las secciones 'Operación N' (por posición), y en
el Resumen se rellenan los importes de 'I. PAGOS A REALIZAR'. La categoría
'Otros' (sin categoría) NO se incluye.

La plantilla tiene 19 columnas (col 1 = PROVEEDOR, col 2 = TIPO, ...). En la
salida se INSERTA una columna 'RUC' en la posición 2 (entre PROVEEDOR y TIPO) y
se OMITE 'N° Registro' (col 18 de la plantilla), porque SUSTENTO ya muestra ese
mismo número con el hipervínculo al PDF. El mapeo src -> dst lo hace `_nc()`:
col 1 -> 1; 2..17 -> +1; 18 -> None (suprimida); 19 -> 19. La salida queda con
19 columnas. Las fórmulas se trasladan con Translator (fila y columna) al
copiar; como ninguna referencia apunta a la columna A, el desplazamiento es
correcto.

Columnas calculadas (supuestos):
- % DET = columna DETRACCION (tasa).  DET = IMPORTE * % DET / 100.  Neto = SALDO - DET - RET.
"""

import re
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from app.services import sharepoint

# --- Columnas de la SALIDA (dst) ---
# Respecto de la plantilla: se inserta 'RUC' en la 2, el 'Neto' se desdobla en
# dos columnas (soles y dólares) y se omite 'N° Registro'.
_COL_RUC = 2
_COL_NETO_SOL = 16
_COL_NETO_USD = 17
# Columna SUSTENTO / LINK FACTURA (donde va el hipervínculo al PDF).
_COL_LINK = 20
_LINK_FONT = Font(color="0563C1", underline="single")

# Columna DET: formato contable con 2 decimales (cero -> guion).
_COL_DET = 13
_DET_FMT = "_-* #,##0.00_-;\\-* #,##0.00_-;_-* \\-??_-;_-@_-"

# Porcentaje con guion en el cero: "12 %" / "-" (para %DET y %RET).
_PCT_FMT = '0 %;-0 %;"-"'

_PLANTILLA = Path(__file__).resolve().parent.parent / "resources" / "plantilla.xlsx"
_DATETIME_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})[ T]\d{2}:\d{2}:\d{2}")
_OPERACION_RE = re.compile(r"^\s*Operaci.n\s+(\d+)", re.IGNORECASE)

# Aseguradoras con las que se trabaja hoy (sección 'PAGOS SEGUROS'). La
# plantilla trae una lista más larga; aquí se reemplaza por estas filas.
_SEGUROS_PROVEEDORES = [
    "RIMAC S.A. ENTIDAD PRESTADORA DE SALUD",
    "RIMAC SEGUROS Y REASEGUROS",
]

# Detalle (SALIDA): columna (1-based) -> clave de texto en los datos.
_TXT = {
    1: "PROVEEDOR", 2: "RUC", 3: "TIPO", 4: "NUMERO",
    5: "FEC REGISTRO", 6: "FECHA DOC.", 7: "FEC. VCTO",
    18: "PRODUCTO", 19: "ORD_COMPRA", 20: "REGISTRO",
}
_FECHA_COLS = {5, 6, 7}


def _num(value) -> float:
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


_FECHA_FMT = "yyyy-mm-dd"

# Plazos de crédito establecidos (días). La resta FEC.VCTO - FEC.DOC se ajusta
# al más cercano de esta lista (p. ej. 6 -> 7, 31 -> 30, 14 -> 15).
_PLAZOS = [1, 7, 15, 30, 45, 60]
# Puntos de corte: 0 y el punto medio entre cada par consecutivo (redondeado
# hacia arriba), para que LOOKUP devuelva el plazo más cercano.
_PLAZO_CORTES = [0] + [
    (a + b + 1) // 2 for a, b in zip(_PLAZOS, _PLAZOS[1:])
]
_PLAZO_LOOKUP = (
    "{" + ",".join(str(c) for c in _PLAZO_CORTES) + "},"
    "{" + ",".join(str(p) for p in _PLAZOS) + "}"
)


def _fecha(value):
    """Fecha como objeto `date` para que Excel pueda operarla (p. ej. PLAZO).
    Si el valor no es una fecha reconocible, devuelve el texto tal cual."""
    s = "" if value is None else str(value).strip()
    m = _DATETIME_RE.match(s)
    if m:
        s = m.group(1)
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return s


def _norm_ruc(ruc) -> str:
    return re.sub(r"\.0$", "", str(ruc).strip())


def _es_ruc_nacional(ruc) -> bool:
    """RUC nacional: 11 dígitos que empiezan con 10 o 20 (persona natural/jurídica)."""
    r = _norm_ruc(ruc)
    return len(r) == 11 and r.isdigit() and r[:2] in ("10", "20")


def _fmt_tipo(v) -> str:
    """Normaliza el TIPO de comprobante: un solo dígito se rellena a 2 ('1' -> '01')."""
    s = re.sub(r"\.0$", "", str(v).strip())
    return "0" + s if s.isdigit() and len(s) == 1 else s


def _key_prov(f) -> str:
    """Clave para ordenar filas por PROVEEDOR (alfabético, sin distinguir may/min)."""
    return str(f.get("PROVEEDOR", "")).strip().upper()


def _moneda_clave(f) -> str:
    """'SOL' o 'USD' (todo lo que no sea soles se trata como dólares)."""
    return "SOL" if str(f.get("MONEDA", "")).strip().upper() == "SOL" else "USD"


# Retención de IGV (SUNAT): 3% del IMPORTE en facturas de bienes que superan
# S/ 700 (o su equivalente en dólares). Valores fijos por norma.
_TASA_RETENCION = 3.0
_UMBRAL_RETENCION = 700.0

# Tipo de cambio (S/ por US$) por defecto y su celda en el Resumen.
TIPO_CAMBIO_DEFAULT = 3.5
_CELDA_TIPO_CAMBIO = "C18"


def _pct_retencion(f: dict, ret_cfg: dict | None) -> float:
    """% de retención de una factura (0 si no aplica).

    - Operación sin retención (p. ej. Pagos servicios) -> 0.
    - Proveedor del exterior (RUC no nacional) -> 0: la retención es a bienes
      nacionales; las importaciones (p. ej. Materia Prima Exterior) no aplican.
    - Con detracción (%DET > 0) -> 0: es un servicio, no un bien.
    - Proveedor que es agente de retención -> 0 (excepción configurada).
    - IMPORTE (convertido a soles) que no supera S/ 700 -> 0.
    - En otro caso -> 3%.
    """
    if not ret_cfg or not ret_cfg.get("activo"):
        return 0.0
    if f.get("__pos") in (ret_cfg.get("pos_sin_ret") or set()):
        return 0.0
    if not _es_ruc_nacional(f.get("RUC", "")):
        return 0.0
    if _num(f.get("DETRACCION")) > 0:
        return 0.0
    if _norm_ruc(f.get("RUC", "")) in (ret_cfg.get("rucs") or set()):
        return 0.0
    importe = _num(f.get("IMPORTE"))
    if str(f.get("MONEDA", "")).strip().upper() != "SOL":
        importe *= float(ret_cfg.get("tipo_cambio") or 0)
    return _TASA_RETENCION if importe > _UMBRAL_RETENCION else 0.0


def _clonar_estilo(d, s) -> None:
    if s.has_style:
        d._style = copy(s._style)


def _quitar_negrita(cell) -> None:
    """Quita la negrita de una celda conservando el resto de la fuente."""
    f = cell.font
    if f is not None and f.bold:
        cell.font = Font(
            name=f.name, size=f.size, bold=False, italic=f.italic,
            vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
            color=f.color,
        )


def _valores_fila(f: dict, ret_cfg: dict | None = None) -> dict:
    """Valores por columna (SALIDA, 1-based) para una fila del Detalle."""
    importe = round(_num(f.get("IMPORTE")), 2)
    pagado = round(_num(f.get("PAGADO")), 2)
    saldo = round(_num(f.get("SALDO")), 2)
    p_det = _num(f.get("DETRACCION"))
    det = round(importe * p_det / 100, 2)
    vals = {i: (_fecha(f.get(k)) if i in _FECHA_COLS else f.get(k, "")) for i, k in _TXT.items()}
    vals[_COL_RUC] = _norm_ruc(f.get("RUC", ""))
    vals[3] = _fmt_tipo(f.get("TIPO", ""))  # TIPO: "1" -> "01"
    # %DET y %RET como fracción (12 -> 0.12) para que Excel muestre "12 %".
    # RET y Neto se escriben como fórmulas (ver _escribir_fila).
    vals.update({8: importe, 9: pagado, 10: saldo, 12: p_det / 100, 13: det})
    vals[14] = _pct_retencion(f, ret_cfg) / 100  # %RET
    return vals


def _neto(f: dict, ret_cfg: dict | None = None) -> float:
    """Neto numérico de una fila, replicando la fórmula del Detalle:
      si DET>0 y |PAGADO-DET|<1  -> SALDO
      si DET>0 y PAGADO=0        -> SALDO-DET
      en otro caso               -> SALDO
    y luego se le resta la retención (RET = %RET * IMPORTE).
    """
    importe = round(_num(f.get("IMPORTE")), 2)
    saldo = round(_num(f.get("SALDO")), 2)
    pagado = round(_num(f.get("PAGADO")), 2)
    det = round(importe * _num(f.get("DETRACCION")) / 100, 2)
    if det > 0 and abs(pagado - det) < 1:
        base = saldo
    elif det > 0 and pagado == 0:
        base = saldo - det
    else:
        base = saldo
    ret = round(importe * _pct_retencion(f, ret_cfg) / 100, 2)
    return round(base - ret, 2)


# Marcador cuando la O/C se consolida por un proveedor relacionado / TIPO 21
# pero no hay una factura del agente en los datos (nombre a mano).
_AGENTE_MANUAL = "Colocar nombre de agente manualmente"

# Tipo de comprobante que, por sí solo, marca una factura como relacionada a un
# agente (va a 'Detalle de agentes' aunque no haya coincidencia de O/C).
_TIPO_AGENTE = "21"


def _tipo(f: dict) -> str:
    return re.sub(r"\.0$", "", str(f.get("TIPO", "")).strip())


def _es_fila_agente(f: dict, ocs_consolidadas: set) -> bool:
    """Una factura va a 'Detalle de agentes' si su O/C está consolidada (tiene un
    agente/proveedor relacionado) o si su TIPO es de agente (21). Si el usuario
    la reasignó manualmente (`__manual`), respeta esa decisión y no va a agentes."""
    if f.get("__manual"):
        return False
    oc = str(f.get("ORD_COMPRA", "")).strip()
    return (bool(oc) and oc in ocs_consolidadas) or _tipo(f) == _TIPO_AGENTE


def _agrupar_agentes(
    filas: list[dict], agente_rucs: list[str], relacionados_rucs: list[str] = None
) -> tuple:
    """Agrupa las facturas que van a 'Detalle de agentes'.

    - Una O/C se consolida (TODAS sus facturas) si incluye un RUC de agente o de
      proveedor relacionado.
    - Además, cualquier factura con TIPO 21 va a agentes por sí sola (aunque su
      O/C no se consolide).

    Devuelve:
      - ocs_consolidadas: set de O/C consolidadas por RUC (para excluir de las
        operaciones normales).
      - nombre_por_oc: O/C -> nombre del agente (o marcador si no hay agente real).
      - ruc_por_oc: O/C -> RUC del agente ("" si no hay agente real).
      - grupos: (O/C, MONEDA) -> lista de filas (O/C "" = facturas TIPO 21 sin O/C).
    """
    agentes = {_norm_ruc(r) for r in (agente_rucs or []) if str(r).strip()}
    relacionados = {_norm_ruc(r) for r in (relacionados_rucs or []) if str(r).strip()}
    disparadores = agentes | relacionados

    # O/C consolidadas (toda la orden) + nombre/ruc del agente real por O/C.
    ocs_consolidadas: set[str] = set()
    agente_nombre_oc: dict[str, str] = {}
    agente_ruc_oc: dict[str, str] = {}
    for f in filas:
        oc = str(f.get("ORD_COMPRA", "")).strip()
        if not oc:
            continue
        ruc = _norm_ruc(f.get("RUC", ""))
        if ruc in disparadores:
            ocs_consolidadas.add(oc)
        if ruc in agentes and oc not in agente_nombre_oc:
            agente_nombre_oc[oc] = str(f.get("PROVEEDOR", "")).strip()
            agente_ruc_oc[oc] = ruc

    # Agrupar todas las filas que van a agentes (por O/C+moneda; sin O/C -> "").
    grupos: dict = {}
    for f in filas:
        if not _es_fila_agente(f, ocs_consolidadas):
            continue
        oc = str(f.get("ORD_COMPRA", "")).strip()
        moneda = str(f.get("MONEDA", "")).strip().upper()
        grupos.setdefault((oc, moneda), []).append(f)

    # Nombre/RUC por O/C: agente real si existe, si no marcador para llenar a mano.
    nombre_por_oc: dict[str, str] = {}
    ruc_por_oc: dict[str, str] = {}
    for oc, _moneda in grupos:
        if oc in agente_nombre_oc:
            nombre_por_oc[oc] = agente_nombre_oc[oc]
            ruc_por_oc[oc] = agente_ruc_oc[oc]
        else:
            nombre_por_oc[oc] = _AGENTE_MANUAL
            ruc_por_oc[oc] = ""

    return ocs_consolidadas, nombre_por_oc, ruc_por_oc, grupos


def _copiar_celda(s, d, src_r: int, dst_r: int, src_col: int, dst_col: int) -> None:
    """Copia estilo y valor de s->d, trasladando fórmulas al nuevo (fila, col)."""
    if s.has_style:
        d._style = copy(s._style)
    v = s.value
    if isinstance(v, str) and v.startswith("=") and (src_r, src_col) != (dst_r, dst_col):
        try:
            v = Translator(
                v, origin=f"{get_column_letter(src_col)}{src_r}"
            ).translate_formula(f"{get_column_letter(dst_col)}{dst_r}")
        except Exception:
            pass  # #REF! u otras fórmulas no trasladables: dejar tal cual
    d.value = v


# Columnas de la plantilla (src) con tratamiento especial.
# 'Neto' se desdobla en dos (soles y dólares); 'N° Registro' se omite porque
# SUSTENTO ya muestra ese mismo número, con el hipervínculo al PDF.
_SRC_COL_NETO = 15
_SRC_COL_OMITIDA = 18


def _nc(c: int) -> int | None:
    """Columna src (plantilla) -> columna dst (salida). None = suprimida."""
    if c == 1:
        return 1
    if c == _SRC_COL_OMITIDA:
        return None
    d = c + 1                          # 'RUC' insertada en la 2
    if c > _SRC_COL_NETO:
        d += 1                         # segunda columna de 'Neto'
    if c > _SRC_COL_OMITIDA:
        d -= 1                         # 'N° Registro' suprimida
    return d


def _nc_rango(c1: int, c2: int) -> tuple[int, int]:
    """Columnas inicial/final (src) de un rango -> (dst). Si un extremo cae en la
    columna suprimida, el rango se encoge hacia el lado que corresponde."""
    return (_nc(c1) or _SRC_COL_OMITIDA + 1), (_nc(c2) or _SRC_COL_OMITIDA)


def _centrar_horizontal(cell) -> None:
    """Centra horizontalmente una celda conservando el resto de la alineación."""
    a = cell.alignment
    cell.alignment = Alignment(
        horizontal="center", vertical=a.vertical, wrap_text=a.wrap_text,
        text_rotation=a.text_rotation, indent=a.indent,
        shrink_to_fit=a.shrink_to_fit,
    )


# Columnas de datos que van centradas horizontalmente (RUC, TIPO, N° DOCUMENTO,
# FEC. REGISTRO, FEC.DOC, FEC.VCTO), mismo layout en Detalle y Detalle de agentes.
_COLS_CENTRAR = range(2, 8)


def _copiar_fila_desplazada(
    src, dst, src_r, dst_r, ncols_src, ruc_val=None, es_cabecera=False
) -> None:
    """Copia una fila de la plantilla a la salida con el desplazamiento de la
    columna RUC. Rellena la col RUC con `ruc_val` (o 'RUC' si es cabecera)."""
    _copiar_celda(src.cell(src_r, 1), dst.cell(dst_r, 1), src_r, dst_r, 1, 1)
    for c in range(2, ncols_src + 1):
        d = _nc(c)
        if d is None:
            continue  # 'N° Registro': no va a la salida
        _copiar_celda(src.cell(src_r, c), dst.cell(dst_r, d), src_r, dst_r, c, d)
    # Columna RUC (2) con el estilo de la columna TIPO (src col 2).
    _clonar_estilo(dst.cell(dst_r, _COL_RUC), src.cell(src_r, 2))
    dst.cell(dst_r, _COL_RUC).value = "RUC" if es_cabecera else ruc_val
    # Segunda columna de 'Neto' (dólares), con el estilo de la primera.
    _clonar_estilo(dst.cell(dst_r, _COL_NETO_USD), dst.cell(dst_r, _COL_NETO_SOL))
    dst.cell(dst_r, _COL_NETO_USD).value = None
    if es_cabecera:
        dst.cell(dst_r, _COL_NETO_SOL).value = "Neto S/"
        dst.cell(dst_r, _COL_NETO_USD).value = "Neto US$"
        # Los encabezados de columna van centrados horizontalmente.
        for c in range(1, _COL_LINK + 1):
            _centrar_horizontal(dst.cell(dst_r, c))


def _copiar_anchos(src, dst) -> None:
    for letra, dim in src.column_dimensions.items():
        if not dim.width:
            continue
        try:
            idx = column_index_from_string(letra)
        except Exception:
            continue
        d = _nc(idx)
        if d is None:
            continue
        dst.column_dimensions[get_column_letter(d)].width = dim.width
    dst.column_dimensions[get_column_letter(_COL_RUC)].width = 16  # RUC


# Anchos de las columnas numéricas del 'Detalle' (para que no salgan "######").
# IMPORTE, PAGADO, SALDO, PLAZO, %DET, DET, %RET, RET, Neto.
_ANCHOS_NUM_DETALLE = {8: 12, 9: 11, 10: 12, 11: 7, 12: 8, 13: 12, 14: 8, 15: 10, 16: 12}


def _detectar_operaciones(ws: Worksheet) -> dict:
    """data_start_row -> (pos, total_row) para cada sección 'Operación N'."""
    secciones: dict = {}
    r = 1
    while r <= ws.max_row:
        a = ws.cell(r, 1).value
        m = _OPERACION_RE.match(str(a)) if a else None
        if m:
            header_row = r + 1
            tr = header_row + 1
            while tr <= ws.max_row and str(ws.cell(tr, 1).value).strip().upper() != "TOTAL":
                tr += 1
            secciones[header_row + 1] = (int(m.group(1)), tr)
            r = tr + 1
        else:
            r += 1
    return secciones


_AGENTE_RE = re.compile(r"AGENTES?\s+DE\s+ADUANAS?\s+(SOL|DOL|D.L)", re.IGNORECASE)


def _detectar_agentes(ws: Worksheet) -> dict:
    """data_start_row -> (moneda, total_row) para 'AGENTES DE ADUANAS SOL/DOL'."""
    secciones: dict = {}
    r = 1
    while r <= ws.max_row:
        a = ws.cell(r, 1).value
        m = _AGENTE_RE.search(str(a)) if a else None
        if m:
            moneda = "SOL" if m.group(1).upper().startswith("SOL") else "USD"
            header_row = r + 1
            tr = header_row + 1
            while tr <= ws.max_row and str(ws.cell(tr, 1).value).strip().upper() != "TOTAL":
                tr += 1
            secciones[header_row + 1] = (moneda, tr)
            r = tr + 1
        else:
            r += 1
    return secciones


_SEGUROS_RE = re.compile(r"PAGOS\s+SEGUROS", re.IGNORECASE)
_PERSONAL_RE = re.compile(r"PAGOS\s+AL\s+PERSONAL", re.IGNORECASE)


def _detectar_por_titulo(ws: Worksheet, regex) -> dict:
    """data_start_row -> total_row de cada sección cuyo título casa con `regex`."""
    secciones: dict = {}
    r = 1
    while r <= ws.max_row:
        a = ws.cell(r, 1).value
        if a and regex.search(str(a)):
            header_row = r + 1
            tr = header_row + 1
            while tr <= ws.max_row and str(ws.cell(tr, 1).value).strip().upper() != "TOTAL":
                tr += 1
            secciones[header_row + 1] = tr
            r = tr + 1
        else:
            r += 1
    return secciones


def _detectar_seguros(ws: Worksheet) -> dict:
    """data_start_row -> total_row para la sección 'PAGOS SEGUROS'."""
    return _detectar_por_titulo(ws, _SEGUROS_RE)


def _detectar_personal(ws: Worksheet) -> dict:
    """data_start_row -> total_row para la sección 'PAGOS AL PERSONAL'."""
    return _detectar_por_titulo(ws, _PERSONAL_RE)


# Columna Neto en la hoja 'Detalle de agentes' (para las fórmulas de enlace).
_COL_NETO_AG = 15


def _ref_agentes(fila: int) -> str:
    """Fórmula que jala el Neto de la hoja 'Detalle de agentes' (columna O)."""
    col = get_column_letter(_COL_NETO_AG)
    return f"=+'Detalle de agentes'!{col}{fila}"


def _escribir_resumen_agente(
    src, estilo_row, dst, r, nombre, ruc, oc, total, ncols_src, ref_row=None,
    moneda="SOL",
):
    """Fila resumen de la sección Agentes: nombre, RUC y O/C del agente y el
    total (Neto) a depositar. Si se da `ref_row`, el total se enlaza por fórmula
    a la hoja 'Detalle de agentes'; si no, se escribe el monto calculado."""
    _copiar_fila_desplazada(src, dst, estilo_row, r, ncols_src, ruc_val=None)
    for c in range(1, _COL_LINK + 1):
        dst.cell(r, c).value = None
    dst.cell(r, 1).value = nombre      # PROVEEDOR
    dst.cell(r, _COL_RUC).value = ruc  # RUC (del agente de la col A)
    # Neto en la columna de su moneda.
    col_neto = (
        _COL_NETO_SOL if str(moneda).strip().upper() == "SOL" else _COL_NETO_USD
    )
    dst.cell(r, col_neto).value = _ref_agentes(ref_row) if ref_row else total
    dst.cell(r, 18).value = nombre     # AGENTE ADUANERO
    dst.cell(r, 19).value = oc         # N° O/C-O/S
    for c in _COLS_CENTRAR:  # RUC y demás columnas de identificación: centradas
        _centrar_horizontal(dst.cell(r, c))
    _centrar_horizontal(dst.cell(r, 19))  # N° O/C-O/S centrado


def _escribir_fila(src, estilo_row, dst, r, fila, ncols_src, sp_cfg, ret_cfg=None) -> None:
    """Escribe una fila de datos del Detalle en `r`, con el estilo (desplazado)
    de `estilo_row`."""
    vals = _valores_fila(fila, ret_cfg)
    _clonar_estilo(dst.cell(r, 1), src.cell(estilo_row, 1))
    dst.cell(r, 1).value = vals.get(1)
    _clonar_estilo(dst.cell(r, _COL_RUC), src.cell(estilo_row, 2))  # RUC (estilo TIPO)
    dst.cell(r, _COL_RUC).value = vals.get(_COL_RUC)
    for c in range(2, ncols_src + 1):
        dc = _nc(c)
        if dc is None:
            continue  # 'N° Registro': no va a la salida
        d = dst.cell(r, dc)
        _clonar_estilo(d, src.cell(estilo_row, c))
        d.value = vals.get(dc)
    for c in _COLS_CENTRAR:  # RUC, TIPO, N° DOC y fechas: centrados
        _centrar_horizontal(dst.cell(r, c))
    _centrar_horizontal(dst.cell(r, 12))  # %DET centrado
    _centrar_horizontal(dst.cell(r, 19))  # N° O/C-O/S centrado
    # Fechas como fecha real (para que PLAZO pueda restarlas).
    for c in _FECHA_COLS:
        if isinstance(vals.get(c), date):
            dst.cell(r, c).number_format = _FECHA_FMT
    # PLAZO (K) = FEC.VCTO (G) - FEC.DOC (F) ajustado al plazo de crédito
    # establecido más cercano; vacío si falta alguna fecha.
    dst.cell(r, 11).value = (
        f'=IF(OR(F{r}="",G{r}=""),"",'
        f'LOOKUP(MAX(0,G{r}-F{r}),{_PLAZO_LOOKUP}))'
    )
    dst.cell(r, 11).number_format = "0"
    _centrar_horizontal(dst.cell(r, 11))
    # DET con dos decimales.
    dst.cell(r, _COL_DET).number_format = _DET_FMT
    # %DET y %RET: porcentaje con guion en el cero (como DET/RET).
    dst.cell(r, 12).number_format = _PCT_FMT
    dst.cell(r, 14).number_format = _PCT_FMT
    # %RET con el mismo borde que %DET (sin el recuadro de "llenar a mano").
    dst.cell(r, 14).border = copy(dst.cell(r, 12).border)
    dst.cell(r, 15).value = f"=ROUND(N{r}*H{r},2)"
    dst.cell(r, 15).number_format = _DET_FMT
    # Neto (fórmula viva): SALDO(J), DET(M), PAGADO(I), RET(O). Va a la columna
    # de su moneda; la otra queda vacía.
    es_sol = str(fila.get("MONEDA", "")).strip().upper() == "SOL"
    col_neto = _COL_NETO_SOL if es_sol else _COL_NETO_USD
    dst.cell(r, col_neto).value = (
        f"=IF(AND(M{r}>0,ABS(I{r}-M{r})<1),J{r},"
        f"IF(AND(M{r}>0,I{r}=0),J{r}-M{r},J{r}))-O{r}"
    )
    # Hipervínculo al PDF en SUSTENTO (nombre del PDF = registro).
    registro = str(vals.get(_COL_LINK) or "").strip()
    if sp_cfg and registro:
        url = sharepoint.link_factura(
            sp_cfg.get("link_principal"), sp_cfg.get("meses"), registro
        )
        if url:
            cel = dst.cell(r, _COL_LINK)
            cel.hyperlink = url
            cel.font = _LINK_FONT
        else:
            dst.cell(r, _COL_LINK).value = None


_MONEDA_TITULO = {"SOL": "Soles", "USD": "Dólares"}


def _titulo_operacion(pos, texto, moneda) -> str:
    """'Operación N - <texto> - <Soles/Dólares>' según la moneda de la config.
    No duplica la moneda si el texto ya la incluye."""
    texto = (texto or "").strip()
    m = str(moneda or "").strip()
    lbl = _MONEDA_TITULO.get(m.upper(), m)
    partes = [f"Operación {pos}"]
    if texto:
        partes.append(texto)
    if lbl and not texto.lower().endswith(lbl.lower()):
        partes.append(lbl)
    return " - ".join(partes)


# Rubros del Detalle (naturaleza del gasto), en el orden en que se emiten.
# `claves`: palabras que se buscan en el nombre de la operación. Las secciones
# fijas de la plantilla se asignan por tipo (ver _RUBRO_POR_TIPO).
_RUBROS = [
    ("MATERIA PRIMA", ("materia prima",)),
    ("LOGÍSTICA Y ADUANAS", ()),
    ("SERVICIOS", ("servicio",)),
    ("PAGOS GENERALES", ("pago masivo", "pagos varios", "pago vario")),
    ("PERSONAL", ()),
    ("SEGUROS", ()),
    ("TESORERÍA", ("transferencia",)),
    ("OTROS PAGOS", ()),
]
_RUBRO_POR_TIPO = {"agentes": 1, "personal": 4, "seguros": 5}
_RUBRO_OTROS = len(_RUBROS) - 1


def _rubro_de(sec: dict, op_texto: dict) -> int:
    """Índice del rubro al que pertenece una sección."""
    if sec["tipo"] in _RUBRO_POR_TIPO:
        return _RUBRO_POR_TIPO[sec["tipo"]]
    texto = str(op_texto.get(sec["pos"]) or "").strip().lower()
    for i, (_nombre, claves) in enumerate(_RUBROS):
        if any(k in texto for k in claves):
            return i
    return _RUBRO_OTROS


def _construir_detalle_sheet(
    wb, grupos, operaciones, fecha_inicio, fecha_final, sp_cfg,
    grupos_agentes=None, nombre_por_oc=None, ruc_por_oc=None, ref_agentes=None,
    ret_cfg=None,
) -> dict:
    src = wb["Detalle"]
    # La plantilla tiene 19 columnas reales (hasta SUSTENTO). La salida tendrá 20
    # (se inserta RUC en la 2).
    ncols = 19
    ops = _detectar_operaciones(src)
    agentes = _detectar_agentes(src)
    seguros = _detectar_seguros(src)
    personal = _detectar_personal(src)
    grupos_agentes = grupos_agentes or {}
    nombre_por_oc = nombre_por_oc or {}
    ruc_por_oc = ruc_por_oc or {}
    ref_agentes = ref_agentes or {"oc": {}, "moneda": {}}
    # Texto/moneda actuales de cada operación (config manda sobre la plantilla).
    op_texto = {o["pos"]: o.get("texto", "") for o in operaciones}
    op_moneda = {o["pos"]: o.get("moneda", "") for o in operaciones}

    dst = wb.create_sheet("__detalle_tmp__")
    _copiar_anchos(src, dst)
    # Anchos fijos para las columnas numéricas (evita "######" en DET, etc.).
    for c, w in _ANCHOS_NUM_DETALLE.items():
        dst.column_dimensions[get_column_letter(c)].width = w

    row_map: dict = {}
    rubros_info: list = []  # un dict por rubro emitido (ver emitir_rubro)
    total_merges: list = []

    # --- Secciones de la plantilla, con sus filas de origen ---
    secciones: list[dict] = []
    for data_row, (pos, total_row) in ops.items():
        secciones.append({"tipo": "operacion", "pos": pos, "moneda": None,
                          "data": data_row, "total": total_row})
    for data_row, (moneda, total_row) in agentes.items():
        secciones.append({"tipo": "agentes", "pos": None, "moneda": moneda,
                          "data": data_row, "total": total_row})
    for data_row, total_row in personal.items():
        secciones.append({"tipo": "personal", "pos": None, "moneda": None,
                          "data": data_row, "total": total_row})
    for data_row, total_row in seguros.items():
        secciones.append({"tipo": "seguros", "pos": None, "moneda": None,
                          "data": data_row, "total": total_row})
    for s in secciones:
        s["titulo"], s["header"] = s["data"] - 2, s["data"] - 1

    # Operaciones creadas en Configuración que la plantilla no tiene: se emiten
    # con el estilo de la última operación de la plantilla.
    plantilla_pos = {p for (p, _tr) in ops.values()}
    if ops:
        modelo_ds = max(ops)
        for pos in sorted(p for p in grupos if p not in plantilla_pos):
            secciones.append({
                "tipo": "operacion", "pos": pos, "moneda": None,
                "data": modelo_ds, "total": ops[modelo_ds][1],
                "titulo": modelo_ds - 2, "header": modelo_ds - 1, "extra": True,
            })

    # --- Agrupación por rubro ---
    por_rubro: dict[int, list] = {}
    for s in secciones:
        por_rubro.setdefault(_rubro_de(s, op_texto), []).append(s)

    primer_titulo = min((s["titulo"] for s in secciones), default=src.max_row + 1)
    # Fila modelo para la banda de rubro: la banda "I. DETALLE DE..." del encabezado.
    banda_modelo = next(
        (r for r in range(1, primer_titulo)
         if isinstance(src.cell(r, 1).value, str)
         and src.cell(r, 1).value.strip().upper().startswith("I.")),
        None,
    )

    def copiar(src_r: int, dst_r: int, *, cabecera=False, mapear=True) -> int:
        """Copia una fila de la plantilla conservando alto y (si toca) merges."""
        _copiar_fila_desplazada(src, dst, src_r, dst_r, ncols, es_cabecera=cabecera)
        if src.row_dimensions[src_r].height:
            dst.row_dimensions[dst_r].height = src.row_dimensions[src_r].height
        if mapear:
            row_map[src_r] = dst_r
        return dst_r + 1

    def emitir_rubro(nombre: str, secs: list, dst_r: int) -> tuple[int, dict]:
        """Un bloque por rubro: banda, una cabecera, todas sus filas (soles y
        dólares juntos) y un TOTAL con las dos monedas."""
        modelo_sec = secs[0]
        tipo = modelo_sec["tipo"]
        modelo, total_row = modelo_sec["data"], modelo_sec["total"]
        alto = src.row_dimensions[modelo].height
        # Las secciones 'extra' reutilizan filas ya emitidas: no deben
        # sobrescribir el mapa de merges.
        mapear = not modelo_sec.get("extra")

        # Banda del rubro (hace de título de la sección).
        if banda_modelo:
            _copiar_fila_desplazada(src, dst, banda_modelo, dst_r, ncols)
            if src.row_dimensions[banda_modelo].height:
                dst.row_dimensions[dst_r].height = src.row_dimensions[banda_modelo].height
        dst.cell(dst_r, 1).value = nombre
        dst.merge_cells(start_row=dst_r, start_column=1, end_row=dst_r,
                        end_column=_COL_LINK)
        dst_r += 1
        dst_r = copiar(modelo_sec["header"], dst_r, cabecera=True, mapear=mapear)

        data_ini = dst_r
        hay = {"SOL": False, "USD": False}
        if tipo == "operacion":
            # Todas las operaciones del rubro, soles primero y luego dólares.
            filas = [f for s in secs for f in grupos.get(s["pos"], [])]
            filas.sort(key=lambda f: (
                str(f.get("MONEDA", "")).strip().upper() != "SOL", _key_prov(f)
            ))
            for f in filas:
                _escribir_fila(src, modelo, dst, dst_r, f, ncols, sp_cfg, ret_cfg)
                if alto:
                    dst.row_dimensions[dst_r].height = alto
                hay[_moneda_clave(f)] = True
                dst_r += 1
        elif tipo == "agentes":
            # Una fila resumen por O/C (de ambas monedas), ordenadas igual.
            resumen = sorted(
                grupos_agentes.items(),
                key=lambda kv: (kv[0][1] != "SOL", kv[0][0]),
            )
            for (oc, mon), fs in resumen:
                _escribir_resumen_agente(
                    src, modelo, dst, dst_r,
                    nombre_por_oc.get(oc, ""), ruc_por_oc.get(oc, ""),
                    oc, round(sum(_neto(f, ret_cfg) for f in fs), 2), ncols,
                    ref_agentes["oc"].get((oc, mon)), mon,
                )
                if alto:
                    dst.row_dimensions[dst_r].height = alto
                hay["SOL" if mon == "SOL" else "USD"] = True
                dst_r += 1
        elif tipo == "seguros":
            # Solo las aseguradoras vigentes (la plantilla trae más).
            for prov in _SEGUROS_PROVEEDORES:
                _copiar_fila_desplazada(src, dst, modelo, dst_r, ncols)
                dst.cell(dst_r, 1).value = prov
                if alto:
                    dst.row_dimensions[dst_r].height = alto
                dst_r += 1
            hay["SOL"] = True
        else:  # personal: filas fijas de la plantilla (bancos)
            for rr in range(modelo, total_row):
                dst_r = copiar(rr, dst_r, mapear=mapear)
            hay["SOL"] = True
        if dst_r == data_ini:  # rubro sin filas: dejar una en blanco
            dst_r = copiar(modelo, dst_r, mapear=mapear)
        data_fin = dst_r - 1

        # TOTAL del rubro, con una suma por moneda.
        fila_total = dst_r
        dst_r = copiar(total_row, dst_r, mapear=False)
        dst.cell(fila_total, 1).value = f"TOTAL {nombre}"
        col_sol = get_column_letter(_COL_NETO_SOL)
        col_usd = get_column_letter(_COL_NETO_USD)
        if tipo == "agentes":
            # Los totales jalan del 'Detalle de agentes' (por moneda).
            for mon, col in (("SOL", _COL_NETO_SOL), ("USD", _COL_NETO_USD)):
                ref = ref_agentes["moneda"].get(mon)
                letra = col_sol if mon == "SOL" else col_usd
                dst.cell(fila_total, col).value = (
                    _ref_agentes(ref) if ref
                    else f"=SUM({letra}{data_ini}:{letra}{data_fin})"
                )
        else:
            dst.cell(fila_total, _COL_NETO_SOL).value = (
                f"=SUM({col_sol}{data_ini}:{col_sol}{data_fin})"
            )
            dst.cell(fila_total, _COL_NETO_USD).value = (
                f"=SUM({col_usd}{data_ini}:{col_usd}{data_fin})"
            )
        total_merges.append(fila_total)
        return dst_r + 1, {  # +1: fila en blanco entre rubros
            "nombre": nombre, "fila": fila_total,
            "sol": hay["SOL"], "usd": hay["USD"],
            "pos": [s["pos"] for s in secs if s["pos"]],
        }

    # --- Emisión: encabezado + un bloque por rubro ---
    dst_r = 1
    for r in range(1, primer_titulo):
        dst_r = copiar(r, dst_r)

    for idx, (nombre_rubro, _claves) in enumerate(_RUBROS):
        secs = por_rubro.get(idx)
        if not secs:
            continue
        dst_r, info = emitir_rubro(nombre_rubro, secs, dst_r)
        rubros_info.append(info)

    # Merges verbatim (de filas copiadas tal cual), con columnas desplazadas.
    for mc in list(src.merged_cells.ranges):
        if mc.min_row in row_map and mc.max_row in row_map:
            dst.merge_cells(
                start_row=row_map[mc.min_row], start_column=_nc_rango(mc.min_col, mc.max_col)[0],
                end_row=row_map[mc.max_row], end_column=_nc_rango(mc.min_col, mc.max_col)[1],
            )
    # Merges de las filas TOTAL de cada rubro (A hasta antes de 'Neto S/').
    for tr in total_merges:
        dst.merge_cells(start_row=tr, start_column=1, end_row=tr,
                        end_column=_COL_NETO_SOL - 1)

    # Título con el rango de fechas.
    rango = ""
    if fecha_inicio or fecha_final:
        rango = f" {fecha_inicio or ''}{' a ' + fecha_final if fecha_final else ''}".rstrip()
    dst.cell(1, 1).value = f"PAGOS PROVEEDORES{rango}"

    # Reemplazar la hoja Detalle por la reconstruida, en la misma posición.
    pos_idx = wb.sheetnames.index("Detalle")
    del wb["Detalle"]
    dst.title = "Detalle"
    wb.move_sheet("Detalle", offset=pos_idx - wb.sheetnames.index("Detalle"))
    return rubros_info


# Banda 'ESTADO DE LIQUIDEZ' del Resumen (cabecera + valores) que se mueve al
# final de la hoja, después de la sección V.
_RESUMEN_BANDA = (4, 5)
_RESUMEN_NCOLS = 4
# Alto de la fila de valores de la banda (para que el texto del estado, que va
# en dos líneas, se lea completo).
_ALTO_BANDA_ESTADO = 32.5
# Tamaños de fuente de esa fila: la plantilla traía 9 / 16 / 17 / 17 (se veían
# desparejos). El texto del estado va menor porque ocupa dos líneas.
_TAM_BANDA_TEXTO = 11
_TAM_BANDA_MONTOS = 16
# Renombrado de rótulos de la banda al moverla.
_RENOMBRE_BANDA = {"TOTAL A PAGAR (US$)": "TOTAL DE VENTA REQUERIDA (US$)"}


def _trasladar_formula(valor, col: str, desde: int, hasta: int):
    """Traslada una fórmula de la fila `desde` a la fila `hasta` (misma columna)."""
    if not (isinstance(valor, str) and valor.startswith("=")) or desde == hasta:
        return valor
    try:
        return Translator(valor, origin=f"{col}{desde}").translate_formula(
            f"{col}{hasta}"
        )
    except Exception:
        return valor


def _desplazar_filas(ws, desde: int, n: int) -> None:
    """Inserta (n>0) o borra (n<0) `n` filas a partir de `desde`.

    openpyxl no ajusta fórmulas, celdas combinadas ni formato condicional al
    mover filas: quedarían desfasados (taparían contenido o pintarían celdas
    equivocadas). Aquí se guardan, se limpian y se rearman con las filas nuevas.
    """
    if not n:
        return
    merges = [
        (m.min_row, m.min_col, m.max_row, m.max_col)
        for m in list(ws.merged_cells.ranges)
    ]
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    cond = [(str(cf.sqref), list(cf.rules)) for cf in ws.conditional_formatting]
    ws.conditional_formatting = ConditionalFormattingList()

    if n > 0:
        ws.insert_rows(desde, n)
    else:
        ws.delete_rows(desde, -n)

    # Las filas que se movieron arrastran sus fórmulas: hay que trasladarlas.
    for row in ws.iter_rows(min_row=desde, max_row=ws.max_row):
        for cel in row:
            v = cel.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            nuevo = _trasladar_formula(
                v, get_column_letter(cel.column), cel.row - n, cel.row
            )
            if nuevo != v:
                cel.value = nuevo

    def _nueva(r: int):
        if r < desde:
            return r                    # arriba de lo movido: sin cambio
        if n < 0 and r < desde - n:
            return None                 # fila borrada
        return r + n

    for min_r, min_c, max_r, max_c in merges:
        f1, f2 = _nueva(min_r), _nueva(max_r)
        if f1 is None or f2 is None or f2 < f1:
            continue
        ws.merge_cells(start_row=f1, start_column=min_c, end_row=f2, end_column=max_c)

    for sqref, reglas in cond:
        rangos = []
        for parte in str(sqref).split():
            cr = CellRange(parte)
            f1, f2 = _nueva(cr.min_row), _nueva(cr.max_row)
            if f1 is None or f2 is None:
                continue
            rangos.append(
                f"{get_column_letter(cr.min_col)}{f1}:"
                f"{get_column_letter(cr.max_col)}{f2}"
            )
        if rangos:
            for regla in reglas:
                ws.conditional_formatting.add(" ".join(rangos), regla)


def _mover_banda_liquidez(ws) -> None:
    """Mueve la banda 'ESTADO DE LIQUIDEZ' (filas 4-5) al final del Resumen y
    ELIMINA las filas 3-5 (la separadora de arriba + la banda), igual que si se
    borraran a mano en Excel.

    Al borrar filas las de abajo suben `n`; openpyxl no ajusta las fórmulas ni
    los merges, así que se trasladan/rearman a mano. Las fórmulas de la banda
    también se ajustan, porque sus referencias (D18, D25, C31, D35-D37) subieron
    lo mismo.

    Debe ejecutarse ANTES de reescribir las fórmulas de 'Operación N' (que
    apuntan a la hoja Detalle), para que esas se escriban ya con la fila final.
    """
    r_ini, r_fin = _RESUMEN_BANDA
    if not str(ws.cell(r_ini, 1).value or "").strip().upper().startswith(
        "ESTADO DE LIQUIDEZ"
    ):
        return  # la plantilla ya no tiene la banda arriba: nada que mover

    r_borrar = r_ini - 1               # fila en blanco que va encima de la banda
    n = (r_fin - r_borrar) + 1         # filas a borrar: separadora + banda

    banda = [
        {
            "row": r,
            "alto": ws.row_dimensions[r].height,
            "celdas": [
                (
                    ws.cell(r, c).value,
                    copy(ws.cell(r, c)._style) if ws.cell(r, c).has_style else None,
                )
                for c in range(1, _RESUMEN_NCOLS + 1)
            ],
        }
        for r in range(r_ini, r_fin + 1)
    ]

    # El formato condicional de la banda (el semáforo) se pierde al borrarla:
    # se guarda para reubicarlo junto con ella.
    cond = [(str(cf.sqref), list(cf.rules)) for cf in ws.conditional_formatting]

    # Borrar la banda y su separador: todo lo de abajo sube n filas.
    _desplazar_filas(ws, r_borrar, -n)

    # Escribir la banda al final, dejando una fila en blanco de separación.
    ultima = max(
        (
            r
            for r in range(1, ws.max_row + 1)
            if any(
                ws.cell(r, c).value not in (None, "")
                for c in range(1, _RESUMEN_NCOLS + 1)
            )
        ),
        default=ws.max_row,
    )
    destino = ultima + 2
    ultima_i = len(banda) - 1
    for i, fila in enumerate(banda):
        r = destino + i
        # La última fila de la banda (los valores) lleva un alto mayor.
        alto = _ALTO_BANDA_ESTADO if i == ultima_i else fila["alto"]
        if alto:
            ws.row_dimensions[r].height = alto
        for c, (valor, estilo) in enumerate(fila["celdas"], start=1):
            cel = ws.cell(r, c)
            if isinstance(valor, str):
                valor = _RENOMBRE_BANDA.get(valor, valor)
            # 'TOTAL DE VENTA REQUERIDA' (col B de la fila de valores): sin
            # fórmula; se llena a mano (queda vacío, con su formato).
            if i == ultima_i and c == 2:
                valor = None
            # Sus referencias también subieron n filas.
            cel.value = _trasladar_formula(
                valor, get_column_letter(c), fila["row"], fila["row"] - n
            )
            if estilo is not None:
                cel._style = estilo
            # Uniformar el tamaño de fuente de la fila de valores (venía disparejo).
            if i == ultima_i:
                f = cel.font
                cel.font = Font(
                    name=f.name,
                    size=_TAM_BANDA_TEXTO if c == 1 else _TAM_BANDA_MONTOS,
                    bold=f.b, italic=f.i, color=f.color,
                    vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
                )

    # Reubicar el formato condicional que vivía DENTRO de la banda (el semáforo);
    # el del resto de la hoja ya lo reacomodó _desplazar_filas.
    for sqref, reglas in cond:
        rangos = []
        for parte in str(sqref).split():
            cr = CellRange(parte)
            if not (r_ini <= cr.min_row <= r_fin and r_ini <= cr.max_row <= r_fin):
                continue
            rangos.append(
                f"{get_column_letter(cr.min_col)}{destino + (cr.min_row - r_ini)}:"
                f"{get_column_letter(cr.max_col)}{destino + (cr.max_row - r_ini)}"
            )
        if rangos:
            for regla in reglas:
                ws.conditional_formatting.add(" ".join(rangos), regla)


# Anchos fijos de columnas del Resumen:
# A = 'Banco/Concepto', B = 'Operación', C = 'Moneda', D = 'Importe'.
_ANCHOS_RESUMEN = {"A": 32, "B": 41.5, "C": 35, "D": 24}


def _ajustar_ancho_operacion(ws) -> None:
    """Fija los anchos de columna del Resumen."""
    for col, ancho in _ANCHOS_RESUMEN.items():
        ws.column_dimensions[col].width = ancho


_BANCO_DEFECTO = "BCP"
_MONEDA_ETIQ_RESUMEN = {"SOL": "S/", "USD": "US$"}


def _fila_etiqueta(ws, texto: str) -> int | None:
    """Fila cuya columna A empieza con `texto` (para ubicar bloques del Resumen)."""
    t = texto.strip().upper()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().upper().startswith(t):
            return r
    return None


def _rellenar_resumen(wb, rubros_info: list, operaciones: list) -> None:
    """Reescribe 'I. PAGOS A REALIZAR' con una fila por rubro y moneda, apuntando
    a los TOTAL del Detalle, y reajusta los totales y el estado de liquidez."""
    if "Resumen" not in wb.sheetnames:
        return
    ws = wb["Resumen"]
    # Primero se reacomoda la hoja (mover banda + borrar filas), y recién luego
    # se escriben las fórmulas hacia Detalle, ya en su fila definitiva.
    _mover_banda_liquidez(ws)

    f_hdr = _fila_etiqueta(ws, "Banco")
    f_tot_sol = _fila_etiqueta(ws, "TOTAL SOLES")
    f_tot_usd = _fila_etiqueta(ws, "TOTAL DÓLARES")
    if not (f_hdr and f_tot_sol):
        return
    f_ini = f_hdr + 1

    # Banco de cada operación, tal como venía en la plantilla (col A).
    op_moneda = {
        o["pos"]: ("SOL" if str(o.get("moneda", "")).strip().upper() == "SOL"
                   else "USD")
        for o in operaciones
    }
    banco_por_pos: dict[int, str] = {}
    for r in range(f_ini, f_tot_sol):
        m = _OPERACION_RE.match(str(ws.cell(r, 2).value or ""))
        if m:
            banco_por_pos[int(m.group(1))] = ws.cell(r, 1).value

    def banco_de(info: dict, mon: str) -> str:
        for pos in info["pos"]:
            if op_moneda.get(pos) == mon and banco_por_pos.get(pos):
                return banco_por_pos[pos]
        return _BANCO_DEFECTO

    # Una fila por rubro y moneda con datos.
    nuevas = [
        (info["nombre"], mon, info["fila"], col, banco_de(info, mon))
        for info in rubros_info
        for mon, col in (("SOL", _COL_NETO_SOL), ("USD", _COL_NETO_USD))
        if info["sol" if mon == "SOL" else "usd"]
    ]
    if not nuevas:
        _ajustar_ancho_operacion(ws)
        return

    # Ajustar cuántas filas hay entre la cabecera y 'TOTAL SOLES'.
    estilo_fila = [
        copy(ws.cell(f_ini, c)._style) if ws.cell(f_ini, c).has_style else None
        for c in range(1, _RESUMEN_NCOLS + 1)
    ]
    diff = len(nuevas) - (f_tot_sol - f_ini)
    if diff:
        _desplazar_filas(ws, f_tot_sol if diff > 0 else f_tot_sol + diff, diff)
        f_tot_sol += diff
        if f_tot_usd:
            f_tot_usd += diff

    # Escribir las filas (banco, rubro, moneda, importe).
    ws.cell(f_hdr, 2).value = "Rubro"
    filas_por_moneda: dict[str, list[int]] = {"SOL": [], "USD": []}
    bancos_usd: dict[str, list[int]] = {}
    for i, (nombre, mon, fila_det, col, banco) in enumerate(nuevas):
        r = f_ini + i
        for c, estilo in enumerate(estilo_fila, start=1):
            if estilo is not None:
                ws.cell(r, c)._style = copy(estilo)
        ws.cell(r, 1).value = banco
        ws.cell(r, 2).value = nombre
        ws.cell(r, 3).value = _MONEDA_ETIQ_RESUMEN[mon]
        ws.cell(r, 4).value = f"=+Detalle!{get_column_letter(col)}{fila_det}"
        filas_por_moneda[mon].append(r)
        if mon == "USD":
            bancos_usd.setdefault(str(banco or _BANCO_DEFECTO).strip().upper(), []).append(r)

    def suma(filas: list[int]) -> str:
        return ("=" + "+".join(f"D{r}" for r in filas)) if filas else 0

    ws.cell(f_tot_sol, 4).value = suma(filas_por_moneda["SOL"])
    if f_tot_usd:
        ws.cell(f_tot_usd, 4).value = suma(filas_por_moneda["USD"])

    # 'IV. ESTADO DE LIQUIDEZ': pagos por cuenta (col B), según banco y moneda.
    f_iv = _fila_etiqueta(ws, "IV.")
    if f_iv:
        for r in range(f_iv, min(f_iv + 8, ws.max_row + 1)):
            etiqueta = str(ws.cell(r, 1).value or "").strip().upper()
            if etiqueta.endswith("SOLES"):
                ws.cell(r, 2).value = f"=+D{f_tot_sol}"
            elif etiqueta.endswith("DÓLARES") or etiqueta.endswith("DOLARES"):
                banco = etiqueta.rsplit(" ", 1)[0].strip()
                ws.cell(r, 2).value = suma(bancos_usd.get(banco, []))

    # La columna de 'Rubro' se ajusta a las etiquetas ya reescritas.
    _ajustar_ancho_operacion(ws)


# Hoja 'Detalle de agentes' (SALIDA): columna -> clave de texto. Layout propio
# (SIN la columna PLAZO del Detalle) con la columna RUC insertada en la 2:
# 1 PROV, 2 RUC, 3 TIPO, 4 NUMERO, 5-7 fechas, 8 IMPORTE, 9 PAGADO, 10 SALDO,
# 11 %DET, 12 DET, 13 %RET, 14 RET, 15 Neto, 16 PRODUCTO, 17 AGENTE, 18 O/C, 20 LINK.
_TXT_AG = {
    1: "PROVEEDOR", 2: "RUC", 3: "TIPO", 4: "NUMERO",
    5: "FEC REGISTRO", 6: "FECHA DOC.", 7: "FEC. VCTO",
    16: "PRODUCTO", 18: "ORD_COMPRA", 19: "REGISTRO",
}


def _valores_fila_ag(f: dict, ret_cfg: dict | None = None) -> dict:
    """Valores por columna (SALIDA) para una fila de 'Detalle de agentes'."""
    importe = round(_num(f.get("IMPORTE")), 2)
    pagado = round(_num(f.get("PAGADO")), 2)
    saldo = round(_num(f.get("SALDO")), 2)
    p_det = _num(f.get("DETRACCION"))
    det = round(importe * p_det / 100, 2)
    vals = {
        i: (_fecha(f.get(k)) if i in _FECHA_COLS else f.get(k, ""))
        for i, k in _TXT_AG.items()
    }
    vals[_COL_RUC] = _norm_ruc(f.get("RUC", ""))
    vals[3] = _fmt_tipo(f.get("TIPO", ""))  # TIPO: "1" -> "01"
    vals.update({8: importe, 9: pagado, 10: saldo, 11: p_det / 100, 12: det})
    vals[13] = _pct_retencion(f, ret_cfg) / 100  # %RET
    return vals


def _escribir_fila_agente(
    src, estilo_row, dst, r, f, ncols_src, sp_cfg, agente, ret_cfg=None
):
    """Escribe una factura en la hoja 'Detalle de agentes' (con columna RUC)."""
    vals = _valores_fila_ag(f, ret_cfg)
    _clonar_estilo(dst.cell(r, 1), src.cell(estilo_row, 1))
    dst.cell(r, 1).value = vals.get(1)
    _clonar_estilo(dst.cell(r, _COL_RUC), src.cell(estilo_row, 2))  # RUC (estilo TIPO)
    dst.cell(r, _COL_RUC).value = vals.get(_COL_RUC)
    for c in range(2, ncols_src + 1):
        dc = _nc(c)
        if dc is None:
            continue  # 'N° Registro': no va a la salida
        d = dst.cell(r, dc)
        _clonar_estilo(d, src.cell(estilo_row, c))
        d.value = vals.get(dc)
    dst.cell(r, 12).number_format = _DET_FMT                       # DET
    for c in _FECHA_COLS:  # fechas como fecha real
        if isinstance(vals.get(c), date):
            dst.cell(r, c).number_format = _FECHA_FMT
    dst.cell(r, 11).number_format = _PCT_FMT                       # %DET: 0 -> "-"
    dst.cell(r, 13).number_format = _PCT_FMT                       # %RET: 0 -> "-"
    dst.cell(r, 13).border = copy(dst.cell(r, 11).border)         # %RET como %DET
    dst.cell(r, 14).value = f"=ROUND(M{r}*H{r},2)"                 # RET = %RET*IMPORTE (2 dec)
    dst.cell(r, 14).number_format = _DET_FMT
    dst.cell(r, 15).value = f"=J{r}-L{r}-N{r}"                     # Neto = SALDO-DET-RET
    if agente:
        dst.cell(r, 17).value = agente                            # AGENTE ADUANERO
    # Los datos de relleno no van en negrita (la fila modelo de la plantilla la trae).
    for c in range(1, _COL_LINK + 1):
        _quitar_negrita(dst.cell(r, c))
    for c in _COLS_CENTRAR:  # RUC, TIPO, N° DOC y fechas: centrados
        _centrar_horizontal(dst.cell(r, c))
    _centrar_horizontal(dst.cell(r, 11))  # %DET centrado
    _centrar_horizontal(dst.cell(r, 18))  # N° O/C-O/S centrado
    registro = str(f.get("REGISTRO") or "").strip()
    if sp_cfg and registro:
        url = sharepoint.link_factura(
            sp_cfg.get("link_principal"), sp_cfg.get("meses"), registro
        )
        if url:
            cel = dst.cell(r, _COL_LINK)
            cel.hyperlink = url
            cel.font = _LINK_FONT


_MONEDA_ETIQUETA = {"SOL": "SOLES", "USD": "DOLARES"}

# Anchos de la hoja 'Detalle de agentes'.
# - Fijos: columnas numéricas / de fórmula / fechas.
_ANCHOS_FIJOS_AG = {
    2: 16, 3: 6, 5: 13, 6: 13, 7: 13, 8: 12, 9: 11, 10: 12,
    11: 8, 12: 12, 13: 8, 14: 11, 15: 12, 19: 14,
}
# - Auto (por contenido) con (mínimo, máximo): columnas de texto.
_ANCHOS_AUTO_AG = {
    1: (18, 45), 4: (12, 22), 16: (16, 55), 17: (18, 45), 18: (11, 18),
}


def _ajustar_anchos_agentes(dst, max_row: int) -> None:
    """Ajusta los anchos de 'Detalle de agentes' para que los datos se vean
    completos (auto por contenido en texto; fijo en numéricas)."""
    for c, w in _ANCHOS_FIJOS_AG.items():
        dst.column_dimensions[get_column_letter(c)].width = w
    for c, (lo, hi) in _ANCHOS_AUTO_AG.items():
        maxlen = 0
        for r in range(1, max_row + 1):
            v = dst.cell(r, c).value
            if v is None or (isinstance(v, str) and v.startswith("=")):
                continue
            maxlen = max(maxlen, len(str(v)))
        dst.column_dimensions[get_column_letter(c)].width = min(max(maxlen + 2, lo), hi)


def _construir_detalle_agentes_sheet(
    wb, grupos_agentes, nombre_por_oc, sp_cfg, ret_cfg=None
):
    """Reconstruye la hoja 'Detalle de agentes' con el detalle de todas las
    facturas agrupadas por O/C. Se separa por moneda (SOLES primero, luego
    DÓLARES), con un subtotal por O/C y un total por moneda.

    Devuelve las filas (en la columna Neto = O) de cada subtotal para que el
    'Detalle' pueda enlazarlas con fórmulas:
      {"oc": {(oc, moneda): fila}, "moneda": {moneda: fila_total}}
    """
    ref = {"oc": {}, "moneda": {}}
    if "Detalle de agentes" not in wb.sheetnames:
        return ref
    src = wb["Detalle de agentes"]
    ncols = 19
    ncols_dst = _nc(ncols)
    header_row, data_style, subtotal_style = 2, 3, 6

    dst = wb.create_sheet("__agentes_tmp__")
    _copiar_anchos(src, dst)

    # Agrupar por moneda: SOLES primero, DÓLARES después, resto al final.
    por_moneda: dict[str, dict] = {}
    for (oc, moneda), filas in grupos_agentes.items():
        por_moneda.setdefault(moneda, {})[oc] = filas
    orden = ["SOL", "USD"]
    monedas = [m for m in orden if m in por_moneda] + [
        m for m in por_moneda if m not in orden
    ]

    alto = src.row_dimensions[data_style].height
    alto_sub = src.row_dimensions[subtotal_style].height
    dst_r = 1
    for moneda in monedas:
        etiqueta = _MONEDA_ETIQUETA.get(moneda, moneda or "SIN MONEDA")
        # Banda de título de la moneda (estilo de fila TOTAL).
        _copiar_fila_desplazada(src, dst, subtotal_style, dst_r, ncols)
        dst.cell(dst_r, 1).value = f"AGENTES DE ADUANAS {etiqueta}"
        dst.cell(dst_r, 15).value = None
        dst.merge_cells(start_row=dst_r, start_column=1, end_row=dst_r, end_column=ncols_dst)
        dst_r += 1
        # Cabecera.
        _copiar_fila_desplazada(src, dst, header_row, dst_r, ncols, es_cabecera=True)
        if src.row_dimensions[header_row].height:
            dst.row_dimensions[dst_r].height = src.row_dimensions[header_row].height
        dst_r += 1
        # Grupos por O/C (ordenados).
        subtotales: list[int] = []
        for oc in sorted(por_moneda[moneda]):
            filas = por_moneda[moneda][oc]
            nombre = nombre_por_oc.get(oc, "")
            data_ini = dst_r
            for i, f in enumerate(filas):
                _escribir_fila_agente(
                    src, data_style, dst, dst_r, f, ncols, sp_cfg,
                    nombre if i == 0 else None, ret_cfg,
                )
                if alto:
                    dst.row_dimensions[dst_r].height = alto
                dst_r += 1
            data_fin = dst_r - 1
            _copiar_fila_desplazada(src, dst, subtotal_style, dst_r, ncols)
            if alto_sub:
                dst.row_dimensions[dst_r].height = alto_sub
            dst.cell(dst_r, 1).value = (f"TOTAL {oc}").strip()
            dst.cell(dst_r, 15).value = f"=SUM(O{data_ini}:O{data_fin})"
            dst.merge_cells(start_row=dst_r, start_column=1, end_row=dst_r, end_column=14)
            subtotales.append(dst_r)
            ref["oc"][(oc, moneda)] = dst_r
            dst_r += 1
        # Total de la moneda.
        _copiar_fila_desplazada(src, dst, subtotal_style, dst_r, ncols)
        if alto_sub:
            dst.row_dimensions[dst_r].height = alto_sub
        dst.cell(dst_r, 1).value = f"TOTAL {etiqueta}"
        dst.cell(dst_r, 15).value = (
            "=" + "+".join(f"O{r}" for r in subtotales) if subtotales else 0
        )
        dst.merge_cells(start_row=dst_r, start_column=1, end_row=dst_r, end_column=14)
        ref["moneda"][moneda] = dst_r
        dst_r += 2  # línea en blanco de separación entre monedas

    _ajustar_anchos_agentes(dst, dst_r)

    pos_idx = wb.sheetnames.index("Detalle de agentes")
    del wb["Detalle de agentes"]
    dst.title = "Detalle de agentes"
    wb.move_sheet(
        "Detalle de agentes",
        offset=pos_idx - wb.sheetnames.index("Detalle de agentes"),
    )
    return ref


def construir_detalle(
    data: dict,
    fecha_inicio: str | None,
    fecha_final: str | None,
    output_path: Path,
    sharepoint_cfg: dict | None = None,
    agente_rucs: list[str] | None = None,
    relacionados_rucs: list[str] | None = None,
    retencion_cfg: dict | None = None,
    tipo_cambio: float | None = None,
    pos_sin_ret: set | None = None,
) -> Path:
    wb = openpyxl.load_workbook(_PLANTILLA)

    tc = float(tipo_cambio) if tipo_cambio else TIPO_CAMBIO_DEFAULT
    # El mismo tipo de cambio va al Resumen (celda del TOTAL CONSOLIDADO).
    if "Resumen" in wb.sheetnames:
        wb["Resumen"][_CELDA_TIPO_CAMBIO] = tc

    # Operaciones que NO aplican retención (p. ej. Pagos servicios). Se prefiere
    # el set recibido (config actual); si no, se deriva del snapshot del proceso.
    if pos_sin_ret is None:
        pos_sin_ret = {
            o["pos"]
            for o in data.get("operaciones", [])
            if not o.get("aplica_retencion", True)
        }
    # Config de retención: interruptor + RUCs exceptuados + tipo de cambio.
    ret_cfg = {
        "activo": bool((retencion_cfg or {}).get("activo")),
        "rucs": {
            _norm_ruc(r)
            for r in ((retencion_cfg or {}).get("rucs") or [])
            if str(r).strip()
        },
        "tipo_cambio": tc,
        "pos_sin_ret": pos_sin_ret,
    }

    # Agrupar por O/C las facturas que incluyen a un agente o proveedor
    # relacionado. Esas filas salen de su Operación normal (van solo a 'Agentes
    # de Aduanas').
    ocs_consolidadas, nombre_por_oc, ruc_por_oc, grupos_agentes = _agrupar_agentes(
        data["filas"], agente_rucs or [], relacionados_rucs or []
    )

    grupos: dict = {}
    for f in data["filas"]:
        if _es_fila_agente(f, ocs_consolidadas):
            continue  # va a 'Agentes de Aduanas' (O/C consolidada o TIPO 21)
        pos = f.get("__pos")
        if pos is None:  # "Otros" no va al Excel
            continue
        grupos.setdefault(pos, []).append(f)

    operaciones = data.get("operaciones", [])
    # Primero 'Detalle de agentes' (para conocer las filas de sus totales) y
    # luego 'Detalle', que enlaza sus resúmenes con fórmulas a esa hoja.
    ref_agentes = _construir_detalle_agentes_sheet(
        wb, grupos_agentes, nombre_por_oc, sharepoint_cfg, ret_cfg
    )
    rubros_info = _construir_detalle_sheet(
        wb, grupos, operaciones, fecha_inicio, fecha_final, sharepoint_cfg,
        grupos_agentes, nombre_por_oc, ruc_por_oc, ref_agentes, ret_cfg,
    )
    _rellenar_resumen(wb, rubros_info, operaciones)

    # Sin líneas de cuadrícula en ninguna hoja (se hace al final para cubrir
    # también las que se reconstruyen).
    for hoja in wb.worksheets:
        hoja.sheet_view.showGridLines = False

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
