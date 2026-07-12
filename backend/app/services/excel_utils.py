"""Utilidades compartidas para leer/escribir archivos tabulares (xls/xlsx/csv)."""

import io
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ProcesamientoError(Exception):
    """Error de validación al procesar un archivo."""


def read_table(filename: str, content: bytes) -> pd.DataFrame:
    """Lee un .csv/.xlsx/.xls a DataFrame (todo como texto)."""
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        try:
            return pd.read_csv(
                io.BytesIO(content), dtype=str, sep=None, engine="python"
            )
        except UnicodeDecodeError:
            return pd.read_csv(
                io.BytesIO(content),
                dtype=str,
                sep=None,
                engine="python",
                encoding="latin-1",
            )
    if suffix in (".xlsx", ".xls"):
        engine = "openpyxl" if suffix == ".xlsx" else "xlrd"
        return pd.read_excel(io.BytesIO(content), dtype=str, engine=engine)
    raise ProcesamientoError(
        f"Formato de archivo no soportado: {suffix or 'desconocido'}"
    )


def write_xlsx(df: pd.DataFrame, path: Path, sheet_name: str = "Datos") -> None:
    """Escribe un DataFrame a .xlsx con encabezado en negrita y anchos fijos."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for i in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22
