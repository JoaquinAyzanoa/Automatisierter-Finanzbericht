from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.schemas.report import FinanceRow


class ExcelService:
    """Builds Excel workbooks from finance data. No HTTP concerns here."""

    HEADERS = ["Account", "Period", "Amount"]

    def build_report(self, name: str, rows: list[FinanceRow], output_path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Finance Report"

        # Title row
        ws["A1"] = name
        ws["A1"].font = Font(size=14, bold=True)

        # Header row
        header_row = 3
        for col, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)

        # Data rows
        total = 0.0
        for i, row in enumerate(rows, start=header_row + 1):
            ws.cell(row=i, column=1, value=row.account)
            ws.cell(row=i, column=2, value=row.period)
            amount_cell = ws.cell(row=i, column=3, value=row.amount)
            amount_cell.number_format = "#,##0.00"
            total += row.amount

        # Total row
        total_row = header_row + 1 + len(rows)
        ws.cell(row=total_row, column=2, value="Total").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=3, value=total)
        total_cell.font = Font(bold=True)
        total_cell.number_format = "#,##0.00"

        # Auto-ish column widths
        for col in range(1, len(self.HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path
